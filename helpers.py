import pandas as pd
from openpyxl import load_workbook
import calendar
import datetime
import datetime

def get_last_day_of_month(date: datetime.date) -> str:
    """Get the last day of the month in format 'DD/M/YYYY'."""
    last_day = calendar.monthrange(date.year, date.month)[1]
    last_date = datetime.date(date.year, date.month, last_day)
    return last_date.strftime(f"{last_day}/{date.month}/{date.year}")

def read_excel(file_path: str, sheet_name: str, header_row: int = 2) -> pd.DataFrame:
    """Read Excel file and return both values and formulas."""
    
    # Read formulas using openpyxl
    workbook = load_workbook(file_path, data_only=False)
    worksheet = workbook[sheet_name]
    
    # Get header row
    headers = []
    for cell in worksheet[header_row + 1]: 
        headers.append(cell.value)
    
    # Read data with formulas
    data_with_formulas = []
    for row_idx, row in enumerate(worksheet.iter_rows(min_row=header_row + 2, values_only=False)):
        row_data = {}
        for col_idx, cell in enumerate(row):
            if col_idx < len(headers) and headers[col_idx]:
                # If cell has a formula, use the formula; otherwise use the value
                if cell.data_type == 'f':  # Formula
                    row_data[headers[col_idx]] = f"{cell.value}"
                else:
                    row_data[headers[col_idx]] = cell.value
        data_with_formulas.append(row_data)
    
    df_formulas = pd.DataFrame(data_with_formulas)
    
    return df_formulas

def parse_month_year(month_year_str):
    """Parse month/year string in format MM/YY to datetime.date object."""
    try:
        # Split the input by '/'
        parts = month_year_str.split('/')
        if len(parts) != 2:
            raise ValueError("Invalid format")
        
        month = int(parts[0])
        year = int(parts[1])
        
        # Convert 2-digit year to 4-digit year
        if year < 100:
            if year >= 0 and year <= 50:  # Assume 00-50 means 2000-2050
                year += 2000
            else:  # 51-99 means 1951-1999
                year += 1900
        
        # Validate month
        if month < 1 or month > 12:
            raise ValueError("Month must be between 1 and 12")
        
        # Return first day of the month
        return datetime.date(year, month, 1)
    
    except (ValueError, IndexError):
        raise ValueError(f"Invalid month/year format: {month_year_str}. Please use MM/YY format (e.g., 05/24)")


def get_user_input():
    """Get user input for month/year and export format."""
    print("Invoice Generation Tool")
    print("=" * 30)
    
    # Get month/year from user
    while True:
        month_year = input("Enter the month/year for invoice generation (MM/YY format, e.g., 05/24): ").strip()
        try:
            target_date = parse_month_year(month_year)
            break
        except ValueError as e:
            print(f"Error: {e}")
            print("Please try again.")
    
    # Get export format from user
    while True:
        export_format = input("Choose export format (excel/csv): ").strip().lower()
        if export_format in ['excel', 'csv']:
            break
        print("Please enter either 'excel' or 'csv'")
    
    return target_date, export_format


def save_dataframe(df, export_format, target_date):
    """Save DataFrame in the specified format."""
    # Generate filename based on target date and format
    month_year_str = target_date.strftime("%m-%Y")
    
    if export_format == "excel":
        output_file = f"output/result_{month_year_str}.xlsx"
        df.to_excel(output_file, index=False)
    else:  # csv
        output_file = f"output/result_{month_year_str}.csv"
        df.to_csv(output_file, index=False)
    
    return output_file